import streamlit as st
import pandas as pd
import io
from datetime import datetime
from analise_jornada_trabalho import analise_jornada_trabalho
from leitor_docx import ler_docx
from leitor_pdf import ler_pdf
from leitor_ocr import ler_imagem

# Tenta importar openpyxl para exportação
try:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    st.warning("Módulo 'openpyxl' não encontrado. A exportação para Excel não funcionará. "
               "Instale-o com `pip install openpyxl`.")


def to_excel(df, resumo):
    """Cria um arquivo Excel em memória com duas abas: Resumo e Detalhado."""
    try:
        # Tenta carregar o modelo. Se não encontrar, cria um novo
        try:
            workbook = load_workbook(filename='modelo_relatorio.xlsx')
        except FileNotFoundError:
            st.error("Arquivo 'modelo_relatorio.xlsx' não encontrado. O download será de um arquivo simples.")
            df_resumo = pd.DataFrame(resumo.items(), columns=['Métrica', 'Valor'])
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name='Resumo da Análise', index=False)
                df.to_excel(writer, sheet_name='Análise Detalhada', index=False)
            buffer.seek(0)
            return buffer

        # Cria um buffer em memória
        buffer = io.BytesIO()
        workbook.save(buffer)

        # Carrega o buffer em um workbook
        workbook = load_workbook(buffer)

        # Preencher a aba 'Resumo da Análise'
        worksheet_resumo = workbook['Resumo da Análise']
        for row in worksheet_resumo.iter_rows(min_row=2, max_row=worksheet_resumo.max_row):
            for cell in row:
                cell.value = None

        df_resumo = pd.DataFrame(resumo.items(), columns=['Métrica', 'Valor'])
        for r_idx, row in enumerate(dataframe_to_rows(df_resumo, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                worksheet_resumo.cell(row=r_idx, column=c_idx, value=value)

        # Preencher a aba 'Análise Detalhada'
        worksheet_detalhada = workbook['Análise Detalhada']
        for row in worksheet_detalhada.iter_rows(min_row=2, max_row=worksheet_detalhada.max_row):
            for cell in row:
                cell.value = None

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                worksheet_detalhada.cell(row=r_idx, column=c_idx, value=value)

        buffer_final = io.BytesIO()
        workbook.save(buffer_final)
        buffer_final.seek(0)
        return buffer_final

    except Exception as e:
        st.error(f"Erro ao gerar o arquivo Excel: {e}")
        return None


# Variáveis de estado do Streamlit para manter os dados
if 'df_analise' not in st.session_state:
    st.session_state.df_analise = None
if 'df_analise_completo' not in st.session_state:
    st.session_state.df_analise_completo = None
if 'resumo_analise' not in st.session_state:
    st.session_state.resumo_analise = None

# --- UI e Lógica do App ---
st.set_page_config(
    page_title="Analisador de Jornada de Trabalho",
    page_icon="⏰",
    layout="wide"
)
st.title("⏰ Analisador de Jornada de Trabalho")
st.markdown("Analise e calcule a jornada de trabalho e horas extras de forma fácil e rápida.")

# --- Seção de Parâmetros de Cálculo ---
st.subheader("Parâmetros do Cálculo")
col1, col2, col3, col4, col5, col6 = st.columns(6)

with col1:
    salario = st.number_input("Salário (R$)", min_value=0.0, format="%.2f", value=2000.0)
with col2:
    jornada_diaria = st.number_input("Jornada Diária (h)", min_value=1.0, value=8.0)
with col3:
    jornada_semanal = st.number_input("Jornada Semanal (h)", min_value=1.0, value=44.0)
with col4:
    intervalo = st.number_input("Intervalo (h)", min_value=0.0, value=1.0)
with col5:
    horario_inicio = st.text_input("Início (HH:MM)", value="08:00")
with col6:
    horario_fim = st.text_input("Fim (HH:MM)", value="18:00")

st.markdown("---")

# --- Seção de Entrada de Dados ---
st.subheader("Entrada de Dados")
tab1, tab2 = st.tabs(["Carregar Arquivo", "Entrada Manual"])

with tab1:
    uploaded_file = st.file_uploader(
        "Selecione um arquivo de registros de ponto",
        type=['txt', 'docx', 'pdf', 'png', 'jpg', 'jpeg'],
        help="O arquivo deve conter os registros de ponto para análise."
    )
    if uploaded_file:
        with st.spinner(f'Analisando o arquivo {uploaded_file.name}...'):
            file_extension = uploaded_file.name.split('.')[-1]
            try:
                if file_extension == 'pdf':
                    st.session_state.texto_registros = ler_pdf(io.BytesIO(uploaded_file.getvalue()))
                elif file_extension == 'docx':
                    st.session_state.texto_registros = ler_docx(io.BytesIO(uploaded_file.getvalue()))
                elif file_extension in ['png', 'jpg', 'jpeg']:
                    st.session_state.texto_registros = ler_imagem(uploaded_file)
                elif file_extension == 'txt':
                    st.session_state.texto_registros = uploaded_file.getvalue().decode("utf-8")
                else:
                    st.session_state.texto_registros = ""
                    st.error("Tipo de arquivo não suportado.")
            except Exception as e:
                st.session_state.texto_registros = ""
                st.error(f"Erro ao processar o arquivo: {e}")

with tab2:
    st.session_state.texto_registros = st.text_area(
        "Cole os registros de ponto aqui (um por linha):",
        st.session_state.get('texto_registros', ""),
        height=300
    )
    if st.button("Limpar Texto Manual"):
        st.session_state.texto_registros = ""
        st.rerun()

if st.button("Calcular Jornada", type="primary", use_container_width=True):
    if not st.session_state.texto_registros:
        st.error("Por favor, insira os registros de ponto ou carregue um arquivo.")
    else:
        with st.spinner('Realizando os cálculos...'):
            try:
                st.session_state.df_analise, st.session_state.resumo_analise = analise_jornada_trabalho(
                    st.session_state.texto_registros,
                    jornada_diaria,
                    jornada_semanal,
                    intervalo,
                    salario,
                    horario_inicio,
                    horario_fim
                )
                st.session_state.df_analise_completo = st.session_state.df_analise.copy()
                st.success("Análise concluída com sucesso!")
            except Exception as e:
                st.error(f"Erro no cálculo: {e}")

# --- Seção de Resultados ---
if st.session_state.df_analise is not None:
    st.markdown("---")
    st.subheader("Resumo da Análise")
    with st.container(border=True):
        col_resumo1, col_resumo2 = st.columns(2)
        with col_resumo1:
            st.metric("Total de Horas Extras", f"{st.session_state.resumo_analise['Total de Horas Extras']:.2f} h")
            st.metric("Horas Extras Normais", f"{st.session_state.resumo_analise['Horas Extras Normais']:.2f} h")
            st.metric("Horas Extras Atípicas", f"{st.session_state.resumo_analise['Horas Extras Atípicas']:.2f} h")
        with col_resumo2:
            st.metric("Custo Total de Horas Extras",
                      f"R$ {st.session_state.resumo_analise['Custo Total de Horas Extras']:.2f}")
            st.metric("Adicional Noturno", f"R$ {st.session_state.resumo_analise['Adicional Noturno']:.2f}")
            st.metric("Inconsistências", st.session_state.resumo_analise['Inconsistencias'])

    st.markdown("---")
    st.subheader("Relatório Detalhado")

    # Filtros
    with st.expander("Filtros Avançados"):
        col_filtro1, col_filtro2, col_filtro3 = st.columns(3)
        with col_filtro1:
            data_especifica = st.date_input("Data Específica", value=None)
        with col_filtro2:
            st.markdown("---")
            filtro_sabado = st.checkbox("Incluir Sábados")
            filtro_domingo = st.checkbox("Incluir Domingos")
        with col_filtro3:
            st.markdown("---")
            filtro_atipico = st.checkbox("Incluir Dias Atípicos")
            filtro_inconsistencia = st.checkbox("Incluir Inconsistências")

    # Lógica de Filtragem
    df_filtrado = st.session_state.df_analise_completo.copy()
    if data_especifica:
        df_filtrado['Data'] = pd.to_datetime(df_filtrado['Data'], dayfirst=True).dt.date
        df_filtrado = df_filtrado[df_filtrado['Data'] == data_especifica]

    if filtro_sabado:
        df_filtrado = df_filtrado[df_filtrado['Dia da Semana'] == 'Sábado']

    if filtro_domingo:
        df_filtrado = df_filtrado[df_filtrado['Dia da Semana'] == 'Domingo']

    if filtro_atipico:
        df_filtrado = df_filtrado[df_filtrado['Observações'].str.contains('Atípico', case=False, na=False)]

    if filtro_inconsistencia:
        df_filtrado = df_filtrado[df_filtrado['Observações'].str.contains('Inconsistência', case=False, na=False)]

    st.dataframe(df_filtrado, use_container_width=True)

    # Botão de exportação
    st.markdown("---")
    excel_file = to_excel(st.session_state.df_analise_completo, st.session_state.resumo_analise)
    if excel_file:
        st.download_button(
            label="📥 Exportar para Excel",
            data=excel_file,
            file_name="relatorio_jornada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )